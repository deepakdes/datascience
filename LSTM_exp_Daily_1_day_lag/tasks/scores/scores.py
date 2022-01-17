#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import numpy as np
import datetime
import keras
import tensorflow as tf
from keras.preprocessing.sequence import TimeseriesGenerator


# In[2]:


from sklearn import metrics 
from sklearn.metrics import mean_squared_error
from sklearn.metrics import r2_score
from yellowbrick.regressor import PredictionError
from yellowbrick.regressor import ResidualsPlot


# In[3]:


import plotly.graph_objs as go
from powerbiclient import Report, models

import openpyxl
from pathlib import Path 
from openpyxl import load_workbook 


# In[14]:


filename = "predictions.csv"
df = pd.read_csv(filename)
print(df.info())


# In[15]:


actuals = df['Actual']
prediction = df['Prediction']


# In[16]:


actuals


# In[17]:


MSE = metrics.mean_squared_error(actuals,prediction)
R_squared = r2_score(actuals, prediction)
MAE = metrics.mean_absolute_error(actuals,prediction)
RMSE = np.sqrt(metrics.mean_squared_error(actuals, prediction))
MAPE = metrics.mean_absolute_percentage_error(actuals,prediction)
time = datetime.datetime.now()
print("MSE:",MSE)
print("R^2:",R_squared)
print("MAE:",MAE)
print("RMSE:",RMSE)
print("MAPE:",MAPE)
print("Time:",time)


# # Write results to csv file

# In[18]:


filename = Path("scores.xlsx")
file_exists = filename.is_file()

header_list = ['MSE','R_squared','MAE','RMSE','MAPE','Time']

fields_scores = [[MSE, R_squared, MAE, RMSE, MAPE,time]]
scores_df = pd.DataFrame(fields_scores)
#scores_df.to_excel("/home/admin/Models/data/LSTM_UNIVARIATE_DAILY/output/scores.xlsx",header=False, index=False)
if file_exists:
    print("filename exist")
    sheetname= 'Sheet1'
    book = load_workbook(filename)
    with pd.ExcelWriter(filename, engine= 'openpyxl', mode ='a') as writer:
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        #print(ws.max_row)
        scores_df.to_excel(writer,sheet_name = sheetname, startrow = writer.sheets['Sheet1'].max_row, header=False, index=False)
        writer.save()
else:
    print("filename does not exist")
    startrow = 0
    sheetname= 'Sheet1'
     # write out the DataFrame to an ExcelWriter
    scores_df.to_excel(filename, sheet_name=sheetname,startrow = startrow, header=header_list, index=False)
    #writer.save()
    #worksheet = writer.sheets[sheet_name]
    #for sheetname in writer.sheets:
        #scores_df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
#writer.close()


# In[ ]:




